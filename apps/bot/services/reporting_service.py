"""
📋 Automated Reporting Service - Bot Service

Enterprise-grade automated reporting with scheduling,
multi-format export, email delivery, and customizable templates.

This service provides comprehensive reporting capabilities for the AnalyticBot,
including report generation, scheduling, and delivery automation.
"""

import asyncio
import json
import logging
import os
import smtplib
import threading
import time
from collections.abc import Callable
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

try:
    from jinja2 import Environment, FileSystemLoader

    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False
    Environment = None
    FileSystemLoader = None

# Scheduling
SCHEDULE_AVAILABLE = False
schedule = None

try:
    import schedule

    SCHEDULE_AVAILABLE = True
except ImportError:
    # Simple fallback - we'll handle the types at usage sites
    pass

# PDF generation dependencies (optional)
REPORTLAB_AVAILABLE = False
letter = None
SimpleDocTemplate = None
Table = None
TableStyle = None
Paragraph = None
Spacer = None
getSampleStyleSheet = None
colors = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    # Simple fallback - we'll handle the types at usage sites
    pass

# Excel generation dependencies (optional)
OPENPYXL_AVAILABLE = False
openpyxl = None
Fill = None
Font = None

try:
    import openpyxl
    from openpyxl.styles import Fill, Font

    OPENPYXL_AVAILABLE = True
except ImportError:
    # Simple fallback - we'll handle the types at usage sites
    pass

# Excel reporting
try:
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.styles import Fill, Font  # type: ignore[import-untyped]

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None
    Fill = None
    Font = None

logger = logging.getLogger(__name__)


class ReportTemplate:
    """
    📄 Report Template for customizable report generation
    """

    def __init__(self, name: str, template_type: str = "standard"):
        self.name = name
        self.template_type = template_type
        self.sections = []
        self.styling = {
            "title_font_size": 16,
            "header_font_size": 14,
            "body_font_size": 12,
            "primary_color": "#1f77b4",
            "secondary_color": "#ff7f0e",
        }

    def add_section(self, section_type: str, title: str, content: Any, **kwargs):
        """Add a section to the report template"""
        section = {
            "type": section_type,
            "title": title,
            "content": content,
            "options": kwargs,
        }
        self.sections.append(section)

    def set_styling(self, **styling_options):
        """Update report styling options"""
        self.styling.update(styling_options)


class AutomatedReportingSystem:
    """
    📋 Automated Reporting System for AnalyticBot

    Enterprise reporting capabilities:
    - Multi-format reports (PDF, Excel, HTML, JSON)
    - Scheduled report generation
    - Email delivery automation
    - Customizable templates and styling
    - Data visualization integration
    - Report versioning and archiving
    """

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True, parents=True)

        self.templates = {}
        self.scheduled_reports = {}
        self.email_config = {}
        self.report_history = {}

        # Initialize Jinja2 environment if available
        if JINJA2_AVAILABLE and Environment and FileSystemLoader:
            self.jinja_env = Environment(loader=FileSystemLoader("templates"), autoescape=True)
        else:
            self.jinja_env = None

    async def create_report(
        self,
        data: pd.DataFrame,
        template: ReportTemplate,
        output_format: str = "pdf",
        filename: str | None = None,
    ) -> dict[str, Any]:
        """
        📊 Create a comprehensive report from data
        """
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{template.name}_{timestamp}"

            output_path = self.output_dir / f"{filename}.{output_format}"

            if output_format == "pdf":
                result = await self._generate_pdf_report(data, template, output_path)
            elif output_format == "excel":
                result = await self._generate_excel_report(data, template, output_path)
            elif output_format == "html":
                result = await self._generate_html_report(data, template, output_path)
            elif output_format == "json":
                result = await self._generate_json_report(data, template, output_path)
            else:
                raise ValueError(f"Unsupported output format: {output_format}")

            # Store report in history
            report_id = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            self.report_history[report_id] = {
                "template": template.name,
                "format": output_format,
                "filename": str(output_path),
                "created_at": datetime.now().isoformat(),
                "data_records": len(data),
                "file_size": output_path.stat().st_size if output_path.exists() else 0,
            }

            result.update({"report_id": report_id, "output_path": str(output_path)})

            return result

        except Exception as e:
            logger.error(f"Report creation failed: {e}")
            return {"error": str(e)}

    async def _generate_pdf_report(
        self, data: pd.DataFrame, template: ReportTemplate, output_path: Path
    ) -> dict[str, Any]:
        """Generate PDF report"""
        try:
            if not REPORTLAB_AVAILABLE:
                return {"error": "ReportLab not available for PDF generation"}

            pagesize = letter
            doc = SimpleDocTemplate(str(output_path), pagesize=pagesize)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=template.styling["title_font_size"],
                spaceAfter=30,
            )
            story.append(Paragraph(template.name, title_style))
            story.append(Spacer(1, 12))

            # Process each section
            for section in template.sections:
                if section["type"] == "header":
                    header_style = ParagraphStyle(
                        "CustomHeader",
                        parent=styles["Heading2"],
                        fontSize=template.styling["header_font_size"],
                    )
                    story.append(Paragraph(section["title"], header_style))
                    story.append(Spacer(1, 12))

                elif section["type"] == "text":
                    story.append(Paragraph(str(section["content"]), styles["Normal"]))
                    story.append(Spacer(1, 12))

                elif section["type"] == "table":
                    if isinstance(section["content"], pd.DataFrame):
                        # Convert DataFrame to table
                        df_section = section["content"].head(20)  # Limit rows for PDF
                        table_data = [df_section.columns.tolist()] + df_section.values.tolist()

                        table = Table(table_data)
                        table.setStyle(
                            TableStyle(
                                [
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                    ("FONTSIZE", (0, 0), (-1, 0), 14),
                                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                                ]
                            )
                        )

                        story.append(table)
                        story.append(Spacer(1, 12))

                elif section["type"] == "summary":
                    # Generate data summary
                    summary_text = self._generate_data_summary(section["content"])
                    story.append(Paragraph(summary_text, styles["Normal"]))
                    story.append(Spacer(1, 12))

            # Add metadata footer
            metadata = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Records: {len(data)}"
            story.append(Spacer(1, 20))
            story.append(Paragraph(metadata, styles["Normal"]))

            doc.build(story)

            return {
                "status": "success",
                "format": "pdf",
                "sections": len(template.sections),
                "file_size": output_path.stat().st_size,
            }

        except Exception as e:
            logger.error(f"PDF report generation failed: {e}")
            return {"error": str(e)}

    async def _generate_excel_report(
        self, data: pd.DataFrame, template: ReportTemplate, output_path: Path
    ) -> dict[str, Any]:
        """Generate Excel report"""
        try:
            if not OPENPYXL_AVAILABLE:
                return {"error": "OpenPyXL not available for Excel generation"}

            workbook = openpyxl.Workbook()  # type: ignore

            # Main data sheet
            worksheet = workbook.active  # type: ignore
            worksheet.title = "Data"  # type: ignore

            # Write data to worksheet
            for col_num, column_title in enumerate(data.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)  # type: ignore
                cell.value = column_title  # type: ignore
                if Font:
                    cell.font = Font(bold=True)  # type: ignore
                if Fill:
                    from openpyxl.styles import PatternFill  # type: ignore

                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )  # type: ignore

            for row_num, row_data in enumerate(data.iterrows(), 2):
                for col_num, value in enumerate(row_data[1], 1):
                    worksheet.cell(row=row_num, column=col_num, value=value)  # type: ignore

            # Auto-adjust column widths
            for column in worksheet.columns:  # type: ignore
                max_length = 0
                try:
                    column_letter = column[0].column_letter  # type: ignore
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception as e:
                            logger.warning(
                                f"Error processing cell value in column {column_letter}: {e}"
                            )
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width  # type: ignore
                except Exception as e:
                    logger.warning(f"Error adjusting column width: {e}")
                    continue

            # Add summary sheet if numeric data exists
            numeric_data = data.select_dtypes(include=[np.number])
            if len(numeric_data.columns) > 0:
                summary_sheet = workbook.create_sheet("Summary")

                # Write summary statistics
                summary_stats = numeric_data.describe()
                for col_num, column_title in enumerate(
                    ["Statistic"] + summary_stats.columns.tolist(), 1
                ):
                    cell = summary_sheet.cell(row=1, column=col_num)
                    cell.value = column_title
                    if Font:
                        cell.font = Font(bold=True)  # type: ignore

                for row_num, (stat_name, row_data) in enumerate(summary_stats.iterrows(), 2):
                    summary_sheet.cell(row=row_num, column=1, value=stat_name)
                    for col_num, value in enumerate(row_data, 2):
                        summary_sheet.cell(row=row_num, column=col_num, value=float(value))

            workbook.save(str(output_path))

            return {
                "status": "success",
                "format": "excel",
                "sheets": len(workbook.worksheets),
                "file_size": output_path.stat().st_size,
            }

        except Exception as e:
            logger.error(f"Excel report generation failed: {e}")
            return {"error": str(e)}

    async def _generate_html_report(
        self, data: pd.DataFrame, template: ReportTemplate, output_path: Path
    ) -> dict[str, Any]:
        """Generate HTML report"""
        try:
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{template.name}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    h1 {{ color: {template.styling["primary_color"]}; }}
                    h2 {{ color: {template.styling["secondary_color"]}; }}
                    table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                    th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                    th {{ background-color: #f2f2f2; font-weight: bold; }}
                    tr:nth-child(even) {{ background-color: #f9f9f9; }}
                    .summary {{ background-color: #e7f3ff; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                </style>
            </head>
            <body>
                <h1>{template.name}</h1>
                <p>Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
            """

            # Process template sections
            for section in template.sections:
                if section["type"] == "header":
                    html_content += f"<h2>{section['title']}</h2>"
                elif section["type"] == "text":
                    html_content += f"<p>{section['content']}</p>"
                elif section["type"] == "table":
                    if isinstance(section["content"], pd.DataFrame):
                        html_content += section["content"].to_html(classes="data-table")
                elif section["type"] == "summary":
                    summary_text = self._generate_data_summary(section["content"])
                    html_content += f'<div class="summary">{summary_text}</div>'

            # Add data preview
            html_content += "<h2>Data Preview</h2>"
            html_content += data.head(10).to_html(classes="data-table")

            # Add summary statistics for numeric columns
            numeric_data = data.select_dtypes(include=[np.number])
            if len(numeric_data.columns) > 0:
                html_content += "<h2>Summary Statistics</h2>"
                html_content += numeric_data.describe().to_html(classes="data-table")

            html_content += """
                <div style="margin-top: 40px; padding-top: 20px; border-top: 1px solid #ccc; color: #666;">
                    <small>Report generated by AnalyticBot Automated Reporting System</small>
                </div>
            </body>
            </html>
            """

            with open(output_path, "w", encoding="utf-8") as f:
                f.write(html_content)

            return {
                "status": "success",
                "format": "html",
                "sections": len(template.sections),
                "file_size": output_path.stat().st_size,
            }

        except Exception as e:
            logger.error(f"HTML report generation failed: {e}")
            return {"error": str(e)}

    async def _generate_json_report(
        self, data: pd.DataFrame, template: ReportTemplate, output_path: Path
    ) -> dict[str, Any]:
        """Generate JSON report"""
        try:
            report_data = {
                "metadata": {
                    "template_name": template.name,
                    "generated_at": datetime.now().isoformat(),
                    "data_shape": data.shape,
                    "columns": data.columns.tolist(),
                },
                "sections": [],
                "data_summary": {},
                "data": data.to_dict(orient="records"),
            }

            # Process template sections
            for section in template.sections:
                section_data = {
                    "type": section["type"],
                    "title": section["title"],
                    "content": (
                        str(section["content"])
                        if not isinstance(section["content"], dict | list)
                        else section["content"]
                    ),
                }
                report_data["sections"].append(section_data)

            # Add data summary
            report_data["data_summary"] = {
                "total_records": len(data),
                "columns_info": {},
                "missing_data": data.isnull().sum().to_dict(),
            }

            # Column information
            for col in data.columns:
                col_info = {
                    "dtype": str(data[col].dtype),
                    "non_null_count": int(data[col].notna().sum()),
                    "unique_values": int(data[col].nunique()),
                }

                if data[col].dtype in ["int64", "float64"]:
                    col_info.update(
                        {
                            "mean": (
                                float(data[col].mean()) if data[col].notna().sum() > 0 else None
                            ),
                            "std": (
                                float(data[col].std()) if data[col].notna().sum() > 0 else None
                            ),
                            "min": (
                                float(data[col].min()) if data[col].notna().sum() > 0 else None
                            ),
                            "max": (
                                float(data[col].max()) if data[col].notna().sum() > 0 else None
                            ),
                        }
                    )

                report_data["data_summary"]["columns_info"][col] = col_info

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)

            return {
                "status": "success",
                "format": "json",
                "sections": len(template.sections),
                "file_size": output_path.stat().st_size,
            }

        except Exception as e:
            logger.error(f"JSON report generation failed: {e}")
            return {"error": str(e)}

    def _generate_data_summary(self, data: pd.DataFrame) -> str:
        """Generate text summary of data"""
        try:
            summary_parts = []

            summary_parts.append(
                f"Dataset contains {len(data):,} records with {len(data.columns)} columns."
            )

            # Data types summary
            dtype_counts = data.dtypes.value_counts()
            dtype_summary = ", ".join([f"{count} {dtype}" for dtype, count in dtype_counts.items()])
            summary_parts.append(f"Column types: {dtype_summary}.")

            # Missing data summary
            missing_total = data.isnull().sum().sum()
            if missing_total > 0:
                missing_pct = (missing_total / data.size) * 100
                summary_parts.append(f"Missing data: {missing_total:,} cells ({missing_pct:.1f}%).")

            # Numeric data summary
            numeric_cols = data.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                summary_parts.append(
                    f"Numeric columns show ranges from {data[numeric_cols].min().min():.2f} to {data[numeric_cols].max().max():.2f}."
                )

            return " ".join(summary_parts)

        except Exception as e:
            logger.error(f"Data summary generation failed: {e}")
            return "Data summary unavailable."

    async def schedule_report(
        self,
        schedule_name: str,
        data_source: Callable,
        template: ReportTemplate,
        schedule_time: str,
        output_format: str = "pdf",
        email_recipients: list[str] | None = None,
    ) -> dict[str, Any]:
        """Schedule automated report generation"""
        try:
            if not SCHEDULE_AVAILABLE:
                return {"error": "Schedule library not available"}

            def generate_scheduled_report():
                """Function to run scheduled report generation"""
                try:
                    # Get fresh data from data source
                    data = data_source()

                    if not isinstance(data, pd.DataFrame):
                        logger.error(f"Data source for {schedule_name} did not return DataFrame")
                        return

                    # Generate report
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"{schedule_name}_{timestamp}"

                    asyncio.run(self.create_report(data, template, output_format, filename))

                    # Send email if recipients provided
                    if email_recipients and self.email_config:
                        output_path = self.output_dir / f"{filename}.{output_format}"
                        asyncio.run(
                            self._send_report_email(
                                email_recipients,
                                f"Scheduled Report: {schedule_name}",
                                str(output_path),
                            )
                        )

                    logger.info(f"Scheduled report {schedule_name} generated successfully")

                except Exception as e:
                    logger.error(f"Scheduled report {schedule_name} failed: {e}")

            # Schedule the report
            if schedule and SCHEDULE_AVAILABLE:
                if schedule_time.lower() == "daily":
                    schedule.every().day.at("09:00").do(generate_scheduled_report)
                elif schedule_time.lower() == "weekly":
                    schedule.every().monday.at("09:00").do(generate_scheduled_report)
                elif schedule_time.lower() == "monthly":
                    # Use monthly scheduling if available
                    if hasattr(schedule.every(), "month"):
                        schedule.every().month.do(generate_scheduled_report)  # type: ignore[attr-defined]
                    else:
                        # Fallback to daily if month is not available
                        schedule.every().day.at("09:00").do(generate_scheduled_report)
                else:
                    # Custom time (e.g., "10:30")
                    schedule.every().day.at(schedule_time).do(generate_scheduled_report)
            else:
                return {"error": "Scheduling not available"}

            self.scheduled_reports[schedule_name] = {
                "template": template.name,
                "schedule": schedule_time,
                "format": output_format,
                "recipients": email_recipients or [],
                "created_at": datetime.now().isoformat(),
            }

            # Start scheduler in background thread if not already running
            if not hasattr(self, "_scheduler_thread") or not self._scheduler_thread.is_alive():
                self._start_scheduler()

            return {
                "status": "scheduled",
                "schedule_name": schedule_name,
                "next_run": (
                    str(schedule.jobs[-1].next_run) if schedule and schedule.jobs else "Unknown"
                ),
            }

        except Exception as e:
            logger.error(f"Report scheduling failed: {e}")
            return {"error": str(e)}

    def _start_scheduler(self):
        """Start the background scheduler thread"""

        def run_scheduler():
            while True:
                try:
                    if schedule and SCHEDULE_AVAILABLE:
                        schedule.run_pending()
                    time.sleep(60)  # Check every minute
                except Exception as e:
                    logger.error(f"Scheduler error: {e}")
                    time.sleep(60)

        self._scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
        self._scheduler_thread.start()

    def configure_email(self, smtp_server: str, smtp_port: int, username: str, password: str):
        """Configure email settings for automated delivery"""
        self.email_config = {
            "smtp_server": smtp_server,
            "smtp_port": smtp_port,
            "username": username,
            "password": password,
        }

    async def _send_report_email(self, recipients: list[str], subject: str, attachment_path: str):
        """Send report via email"""
        try:
            if not self.email_config:
                raise ValueError("Email not configured")

            msg = MIMEMultipart()
            msg["From"] = self.email_config["username"]
            msg["To"] = ", ".join(recipients)
            msg["Subject"] = subject

            # Email body
            body = f"""
            Automated Report Generated
            
            Please find the attached report generated by AnalyticBot.
            
            Generated at: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
            
            Best regards,
            AnalyticBot Reporting System
            """

            msg.attach(MIMEText(body, "plain"))

            # Attach report file
            if os.path.exists(attachment_path):
                with open(attachment_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())

                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename= {os.path.basename(attachment_path)}",
                )
                msg.attach(part)

            # Send email
            server = smtplib.SMTP(self.email_config["smtp_server"], self.email_config["smtp_port"])
            server.starttls()
            server.login(self.email_config["username"], self.email_config["password"])
            server.sendmail(self.email_config["username"], recipients, msg.as_string())
            server.quit()

            logger.info(f"Report email sent to {recipients}")

        except Exception as e:
            logger.error(f"Email sending failed: {e}")

    def get_report_history(self) -> dict[str, Any]:
        """Get history of generated reports"""
        return self.report_history

    def get_scheduled_reports(self) -> dict[str, Any]:
        """Get list of scheduled reports"""
        return self.scheduled_reports

    async def health_check(self):
        """Health check for the reporting system"""
        return {
            "status": "healthy",
            "reports_generated": len(self.report_history),
            "scheduled_reports": len(self.scheduled_reports),
            "output_directory": str(self.output_dir),
            "dependencies": {
                "reportlab": REPORTLAB_AVAILABLE,
                "openpyxl": OPENPYXL_AVAILABLE,
                "jinja2": JINJA2_AVAILABLE,
                "schedule": SCHEDULE_AVAILABLE,
            },
        }


# Convenience functions for easy integration with bot services
async def create_reporting_system(output_dir: str = "reports"):
    """Factory function to create reporting system"""
    return AutomatedReportingSystem(output_dir=output_dir)


def create_report_template(name: str, template_type: str = "standard"):
    """Factory function to create report template"""
    return ReportTemplate(name=name, template_type=template_type)


# Example usage and testing
if __name__ == "__main__":
    # Create sample data for testing
    np.random.seed(42)

    sample_data = pd.DataFrame(
        {
            "date": pd.date_range("2023-01-01", periods=100, freq="D"),
            "sales": np.random.normal(10000, 2000, 100),
            "customers": np.random.poisson(50, 100),
            "region": np.random.choice(["North", "South", "East", "West"], 100),
            "product": np.random.choice(["A", "B", "C", "D"], 100),
        }
    )

    # Test the reporting system
    reporting_system = AutomatedReportingSystem()

    print("📋 Testing Automated Reporting System...")

    async def test_reporting():
        # Create a sample template
        template = create_report_template("Sales Report", "analytics")
        template.add_section("header", "Executive Summary", "")
        template.add_section(
            "text", "Overview", "This report provides insights into sales performance."
        )
        template.add_section("table", "Sales Data", sample_data.head(10))
        template.add_section("summary", "Data Summary", sample_data)

        # Generate reports in different formats
        pdf_result = await reporting_system.create_report(sample_data, template, "pdf")
        print(f"PDF Report: {pdf_result.get('status', 'failed')}")

        html_result = await reporting_system.create_report(sample_data, template, "html")
        print(f"HTML Report: {html_result.get('status', 'failed')}")

        json_result = await reporting_system.create_report(sample_data, template, "json")
        print(f"JSON Report: {json_result.get('status', 'failed')}")

        # Test health check
        health = await reporting_system.health_check()
        print(f"Health Status: {health['status']}")
        print(f"Reports Generated: {health['reports_generated']}")

    asyncio.run(test_reporting())

    print("✅ Automated Reporting System test complete!")
